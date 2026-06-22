import base64
import hashlib
import json
import os
import re
import tempfile
import threading
import time
import tkinter as tk
import uuid
import sys
from copy import copy
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD

    BaseTk = TkinterDnD.Tk
except ImportError:
    DND_FILES = None
    BaseTk = tk.Tk


MIN_NORMAL = 300
MAX_NORMAL = 600

RED_FONT = Font(color="FF0000")
DEFAULT_FONT = Font(color="000000")
HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")

PEN_TITLE_PATTERN = re.compile(r"(\d+)栏\s*(南|北)")
COUNT_PATTERN = re.compile(r"(\d+)\s*只")
ROOSTER_PATTERN = re.compile(r"公鸡")

HEN_HEADERS = [
    "1栏南",
    "1栏北",
    "2栏南",
    "2栏北",
    "3栏南",
    "3栏北",
    "4栏南",
    "4栏北",
    "5栏南",
    "5栏北",
]
ROOSTER_HEADERS = [f"公鸡{index}" for index in range(1, 6)]
FIXED_HEADERS = HEN_HEADERS + ROOSTER_HEADERS
BUILDING_CHOICES = ["1-1", "1-2", "2-1", "2-2", "3-1", "3-2", "4", "5", "6"]
BUILDING_PATTERN = re.compile(r"(?<!\d)(\d{1,3})\s*[栋舍](?!\d)")
LEADING_BUILDING_PATTERN = re.compile(r"^\s*(\d{1,3})(?!\d)")

EXCEL_EXTENSIONS = {".xlsx", ".xlsm"}
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
SUPPORTED_EXTENSIONS = EXCEL_EXTENSIONS | IMAGE_EXTENSIONS

QUARK_API_URL = "https://scan-business.quark.cn/vision"
QUARK_BUSINESS = "vision"
DEFAULT_SIGN_METHOD = "SHA3-256"
MAX_QUARK_UPLOAD_BYTES = 5 * 1024 * 1024
DEFAULT_CLIENT_ID = (
    os.getenv("QUARK_SCAN_CLIENT_ID")
    or os.getenv("QUARK_SCAN_API_KEY_ID")
    or ""
)
DEFAULT_CLIENT_SECRET = (
    os.getenv("QUARK_SCAN_CLIENT_SECRET")
    or os.getenv("QUARK_SCAN_API_KEY")
    or ""
)
RETRYABLE_CODES = {"A0300"}


def desktop_path():
    return Path.home() / "Desktop"


def building_label(value):
    return f"{value}舍" if value else ""


def make_unique_path(path):
    path = Path(path)
    if not path.exists():
        return path
    for index in range(1, 1000):
        candidate = path.with_name(f"{path.stem}（{index}）{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"无法生成不重复的文件名：{path}")


def safe_sheet_title(title, existing_titles):
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(title).strip()) or "Sheet"
    base = cleaned[:31]
    if base not in existing_titles:
        return base
    for index in range(1, 1000):
        suffix = f"_{index}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        if candidate not in existing_titles:
            return candidate
    raise RuntimeError(f"无法生成不重复的工作表名称：{title}")


def is_image_file(path):
    return Path(path).suffix.lower() in IMAGE_EXTENSIONS


def expand_input_paths(paths):
    expanded = []
    for path in paths:
        file_path = Path(path)
        if file_path.is_dir():
            for child in sorted(file_path.rglob("*")):
                if (
                    child.is_file()
                    and child.suffix.lower() in SUPPORTED_EXTENSIONS
                    and not child.name.startswith("~$")
                ):
                    expanded.append(str(child))
        elif (
            file_path.is_file()
            and file_path.suffix.lower() in SUPPORTED_EXTENSIONS
            and not file_path.name.startswith("~$")
        ):
            expanded.append(str(file_path))
    return expanded


def normalize_title(value):
    if value is None:
        return None
    text = str(value).strip()
    pen_match = PEN_TITLE_PATTERN.search(text)
    if pen_match:
        return f"{pen_match.group(1)}栏{pen_match.group(2)}"
    if ROOSTER_PATTERN.search(text):
        return "公鸡"
    return None


def parse_expected_count(value):
    if value is None:
        return None
    match = COUNT_PATTERN.search(str(value))
    return int(match.group(1)) if match else None


def is_fixed_output_header(value):
    if value is None:
        return False
    return str(value).strip() in FIXED_HEADERS


def source_max_column(ws):
    # If an already-arranged file is loaded, fixed headers may exist on row 1
    # to the right of the source table. Stop before those columns to avoid
    # collecting generated output again.
    for col in range(2, ws.max_column + 1):
        if is_fixed_output_header(ws.cell(row=1, column=col).value):
            return col - 1
    return ws.max_column


def cell_display_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def numeric_display_value(value):
    value = cell_display_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, str) and re.fullmatch(r"\d+(?:\.0)?", value.strip()):
        number = float(value)
        return int(number) if number.is_integer() else number
    return None


def parse_embedded_table(value):
    if not isinstance(value, str):
        return None
    text = value.strip()
    if not text.startswith("{"):
        return None
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return None
    rows = payload.get("list")
    if not isinstance(rows, list):
        return None

    items = []
    for row in rows:
        if not isinstance(row, list):
            continue
        for value in row:
            number = numeric_display_value(value)
            if number is not None:
                items.append({"value": number, "coord": None})
    return items or None


def extract_building_name_from_text(text):
    if not text:
        return None
    text = str(text).strip()
    match = BUILDING_PATTERN.search(text)
    if match:
        return f"{match.group(1)}舍"
    match = LEADING_BUILDING_PATTERN.search(text)
    if match:
        return f"{match.group(1)}舍"
    return None


def extract_building_name_from_workbook(excel_path, fallback_path=None):
    wb = None
    try:
        wb = load_workbook(excel_path, data_only=False)
        ws = wb.active
        for row in range(1, min(ws.max_row, 8) + 1):
            parts = []
            for col in range(1, min(ws.max_column, 8) + 1):
                value = ws.cell(row=row, column=col).value
                if value not in (None, ""):
                    parts.append(str(value))
            name = extract_building_name_from_text(" ".join(parts))
            if name:
                return name
    finally:
        if wb is not None:
            wb.close()

    if fallback_path:
        path = Path(fallback_path)
        return (
            extract_building_name_from_text(path.stem)
            or extract_building_name_from_text(path.parent.name)
        )
    return None


def style_by_value(cell, value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        cell.font = DEFAULT_FONT
        return
    cell.font = RED_FONT if number < MIN_NORMAL or number > MAX_NORMAL else DEFAULT_FONT


def split_numeric_groups(ws):
    groups = []
    current = []
    for row in range(1, ws.max_row + 1):
        values = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            embedded_items = parse_embedded_table(cell.value)
            if embedded_items:
                if current:
                    groups.append(current)
                    current = []
                groups.append(embedded_items)
                continue

            value = numeric_display_value(cell.value)
            if value is not None:
                values.append({"value": value, "coord": cell.coordinate})

        if values:
            current.extend(values)
        elif current:
            groups.append(current)
            current = []

    if current:
        groups.append(current)
    return groups


def extract_data_from_groups(ws):
    groups = split_numeric_groups(ws)
    if len(groups) < 10:
        return None

    data = {header: {"items": [], "expected": None} for header in FIXED_HEADERS}
    for pen in range(1, 6):
        south_index = pen - 1
        north_index = pen + 4
        if south_index < len(groups):
            data[f"{pen}栏南"]["items"].extend(groups[south_index])
            data[f"{pen}栏南"]["expected"] = len(groups[south_index])
        if north_index < len(groups):
            data[f"{pen}栏北"]["items"].extend(groups[north_index])
            data[f"{pen}栏北"]["expected"] = len(groups[north_index])

    for index, group in enumerate(groups[10:15], start=1):
        header = f"公鸡{index}"
        data[header]["items"].extend(group)
        data[header]["expected"] = len(group)

    return data if any(column["items"] for column in data.values()) else None


def get_quark_credentials():
    candidate_files = [Path(__file__).with_name("quark_credentials.json")]
    if getattr(sys, "frozen", False):
        candidate_files.insert(0, Path(sys.executable).with_name("quark_credentials.json"))

    for local_file in candidate_files:
        if not local_file.exists():
            continue
        config = json.loads(local_file.read_text(encoding="utf-8"))
        client_id = config.get("client_id") or config.get("clientId")
        client_secret = config.get("client_secret") or config.get("clientSecret")
        if client_id and client_secret:
            return str(client_id).strip(), str(client_secret).strip()

    client_id = DEFAULT_CLIENT_ID.strip() if DEFAULT_CLIENT_ID else ""
    client_secret = DEFAULT_CLIENT_SECRET.strip() if DEFAULT_CLIENT_SECRET else ""
    return client_id, client_secret


def make_signature(client_id, client_secret, sign_method, sign_nonce, timestamp):
    raw = f"{client_id}_{QUARK_BUSINESS}_{sign_method}_{sign_nonce}_{timestamp}_{client_secret}"
    data = raw.encode("utf-8")
    method = sign_method.lower().replace("_", "-")
    if method == "sha3-256":
        digest = hashlib.sha3_256(data)
    elif method == "sha256":
        digest = hashlib.sha256(data)
    elif method == "sha1":
        digest = hashlib.sha1(data)
    elif method == "md5":
        digest = hashlib.md5(data)
    else:
        raise ValueError("Unsupported sign method: " + sign_method)
    return digest.hexdigest().lower()


def compress_image_for_quark(image_path, temp_dir):
    image_path = Path(image_path)
    if image_path.stat().st_size <= MAX_QUARK_UPLOAD_BYTES:
        return image_path

    output_path = Path(temp_dir) / f"{image_path.stem}_compressed.jpg"
    with Image.open(image_path) as image:
        image = image.convert("RGB")
        max_side = 3000
        if max(image.size) > max_side:
            image.thumbnail((max_side, max_side), Image.LANCZOS)

        for quality in (92, 88, 84, 80, 76, 72, 68, 64, 60):
            image.save(output_path, "JPEG", quality=quality, optimize=True)
            if output_path.stat().st_size <= MAX_QUARK_UPLOAD_BYTES:
                return output_path

    raise RuntimeError(
        f"图片压缩后仍超过 5MB：{image_path.name}，请裁剪或降低照片分辨率后再试。"
    )


def image_to_base64(image_path):
    path = Path(image_path)
    if not path.is_file():
        raise FileNotFoundError("图片文件不存在：" + str(path))
    return base64.b64encode(path.read_bytes()).decode("ascii")


def build_quark_payload(image_path, client_id, client_secret):
    sign_nonce = uuid.uuid4().hex
    timestamp = int(time.time() * 1000)
    sign_method = DEFAULT_SIGN_METHOD
    return {
        "dataType": "image",
        "serviceOption": "typeset",
        "inputConfigs": json.dumps({"function_option": "excel"}, ensure_ascii=False),
        "outputConfigs": json.dumps({"need_return_image": "False"}, ensure_ascii=False),
        "dataBase64": image_to_base64(image_path),
        "reqId": uuid.uuid4().hex,
        "clientId": client_id,
        "signMethod": sign_method,
        "signNonce": sign_nonce,
        "timestamp": timestamp,
        "signature": make_signature(
            client_id, client_secret, sign_method, sign_nonce, timestamp
        ),
    }


def extract_excel_base64(response_json):
    data = response_json.get("data") or {}
    typeset_info = data.get("TypesetInfo")
    if not typeset_info:
        raise RuntimeError("夸克接口未返回 TypesetInfo：" + str(response_json))

    for item in typeset_info:
        if isinstance(item, dict) and str(item.get("FileType", "")).lower() == "excel":
            return item.get("FileBase64")

    first = typeset_info[0] if typeset_info else {}
    if isinstance(first, dict):
        return first.get("FileBase64")
    raise RuntimeError("夸克接口未返回 Excel 数据：" + str(typeset_info))


def convert_image_to_excel(image_path, output_path, temp_dir=None, max_retries=3):
    client_id, client_secret = get_quark_credentials()
    if not client_id or not client_secret:
        raise RuntimeError(
            "未配置夸克 API Key ID / API Key。请设置环境变量 "
            "QUARK_SCAN_CLIENT_ID 和 QUARK_SCAN_CLIENT_SECRET，"
            "或在同目录 quark_credentials.json 中填写。"
        )

    temp_dir = temp_dir or tempfile.gettempdir()
    upload_image = compress_image_for_quark(image_path, temp_dir)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    last_error = None
    for attempt in range(1, max_retries + 1):
        payload = build_quark_payload(upload_image, client_id, client_secret)
        try:
            response = requests.post(
                QUARK_API_URL,
                json=payload,
                headers={"Content-Type": "application/json"},
                timeout=180,
                allow_redirects=True,
            )
            response.raise_for_status()
            body = response.json()
        except requests.RequestException as exc:
            last_error = exc
            if attempt < max_retries:
                time.sleep(min(2 * attempt, 6))
                continue
            raise RuntimeError(f"夸克接口请求失败，已重试 {attempt} 次：{exc}") from exc
        except ValueError as exc:
            raise RuntimeError(
                "夸克接口返回内容不是 JSON，可能是鉴权参数或接口地址不匹配。"
            ) from exc

        code = body.get("code")
        if code not in (0, "0", "00000", None):
            message = body.get("message") or body.get("msg") or "UNKNOWN"
            req_id = body.get("reqId") or body.get("req_id") or payload.get("reqId")
            last_error = RuntimeError(
                f"夸克接口错误：code={code}, message={message}, reqId={req_id}"
            )
            if str(code) in RETRYABLE_CODES and attempt < max_retries:
                time.sleep(min(4 * attempt, 12))
                continue
            raise last_error

        excel_base64 = extract_excel_base64(body)
        if not excel_base64:
            raise RuntimeError("夸克接口返回中没有 Excel 文件数据。")

        output_path.write_bytes(base64.b64decode(excel_base64))
        return output_path

    raise RuntimeError(f"夸克图片转 Excel 失败：{last_error}")


def extract_arranged_data(excel_path):
    wb = None
    try:
        wb = load_workbook(excel_path, data_only=False)
        ws = wb.active
        if ws.max_row < 2 or ws.max_column < 1:
            raise RuntimeError("Excel 工作表数据过少。")

        max_col = source_max_column(ws)
        data = {header: {"items": [], "expected": None} for header in FIXED_HEADERS}
        current_title = None
        rooster_index = 0

        for row in range(1, ws.max_row + 1):
            title_cell = ws.cell(row=row, column=1)
            title = normalize_title(title_cell.value)
            if title:
                if title == "公鸡":
                    rooster_index += 1
                    current_title = f"公鸡{rooster_index}" if rooster_index <= 5 else None
                else:
                    current_title = title if title in data else None
                if current_title:
                    data[current_title]["expected"] = parse_expected_count(title_cell.value)
                continue

            if not current_title:
                continue

            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                value = cell_display_value(cell.value)
                if value is not None:
                    data[current_title]["items"].append(
                        {"value": value, "coord": cell.coordinate}
                    )

        grouped_data = extract_data_from_groups(ws)
        data_total = sum(len(column["items"]) for column in data.values())
        grouped_total = (
            sum(len(column["items"]) for column in grouped_data.values())
            if grouped_data
            else 0
        )
        data_complete = all(data[header]["items"] for header in FIXED_HEADERS)

        if grouped_data and (not data_complete or grouped_total > data_total):
            return grouped_data

        if not any(column["items"] for column in data.values()):
            raise RuntimeError(
                "未识别到任何栏位数据。请确认识别结果包含“1栏南/1栏北/公鸡”等标题，"
                "或包含 10 个以上按空行分隔的数字块。"
            )
        return data
    finally:
        if wb is not None:
            wb.close()


def copy_source_sheet(source_ws, target_ws):
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged_range))

    for key, dimension in source_ws.column_dimensions.items():
        if dimension.width:
            target_ws.column_dimensions[key].width = dimension.width

    for key, dimension in source_ws.row_dimensions.items():
        if dimension.height:
            target_ws.row_dimensions[key].height = dimension.height


def arranged_header_text(header, column_data):
    expected = column_data.get("expected")
    actual = len(column_data.get("items", []))
    count = expected if expected is not None else actual
    return f"{header}（{count}只）" if count else header


def write_arranged_sheet(output_wb, sheet_name, excel_path, data):
    ws = output_wb.create_sheet(safe_sheet_title(sheet_name, output_wb.sheetnames))
    source_wb = load_workbook(excel_path, data_only=False)
    try:
        source_ws = source_wb.active
        copy_source_sheet(source_ws, ws)

        start_col = source_ws.max_column + 2
        max_rows = max((len(column["items"]) for column in data.values()), default=0)

        for offset, header in enumerate(FIXED_HEADERS):
            col = start_col + offset
            column_data = data.get(header, {"items": [], "expected": None})
            cell = ws.cell(row=1, column=col)
            cell.value = arranged_header_text(header, column_data)
            cell.font = Font(bold=True, color="000000")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(str(cell.value)) + 2)

            for row, item in enumerate(column_data["items"], start=2):
                value_cell = ws.cell(row=row, column=col)
                value_cell.value = f"={item['coord']}" if item.get("coord") else item["value"]
                style_by_value(value_cell, item["value"])

            actual = len(column_data["items"])
            expected = column_data.get("expected")
            if expected is None:
                expected = actual

            count_row = max_rows + 3
            check_row = max_rows + 4
            count_cell = ws.cell(row=count_row, column=col)
            count_cell.value = f"识别数量：{actual}/{expected}"
            count_cell.font = Font(bold=True, color="000000")
            count_cell.alignment = Alignment(horizontal="center")

            check_cell = ws.cell(row=check_row, column=col)
            check_cell.value = "核对无误" if actual == expected else f"数量不符：应{expected}实{actual}"
            check_cell.font = Font(
                bold=True,
                color="008000" if actual == expected else "FF0000",
            )
            check_cell.alignment = Alignment(horizontal="center")

        return max_rows
    finally:
        source_wb.close()


def prepare_excel_from_input(input_path, temp_dir, progress_callback=None):
    input_path = Path(input_path)
    suffix = input_path.suffix.lower()

    if suffix in IMAGE_EXTENSIONS:
        if progress_callback:
            progress_callback(20, "正在调用夸克图片转 Excel...")
        temp_excel = Path(temp_dir) / f"{input_path.stem}_quark.xlsx"
        convert_image_to_excel(input_path, temp_excel, temp_dir=temp_dir)
        if progress_callback:
            progress_callback(65, "图片已转为 Excel，正在整理标签页...")
        return temp_excel

    if suffix in EXCEL_EXTENSIONS:
        if progress_callback:
            progress_callback(45, "正在整理 Excel 标签页...")
        return input_path

    raise RuntimeError("请选择图片文件或 .xlsx/.xlsm 文件。")


def rename_image_to_building(input_path, building_name):
    input_path = Path(input_path)
    if input_path.suffix.lower() not in IMAGE_EXTENSIONS or not building_name:
        return input_path
    target = make_unique_path(input_path.with_name(f"{building_name}{input_path.suffix.lower()}"))
    if target == input_path:
        return input_path
    input_path.rename(target)
    return target


def default_output_name(save_path):
    return Path(save_path).stem in {"称重单识别整理结果", "识别整理结果", "整理结果"}


def process_inputs_to_workbook(
    input_paths,
    save_path,
    progress_callback=None,
    auto_rename=True,
    selected_building=None,
):
    input_paths = [Path(path) for path in input_paths]
    save_path = Path(save_path)
    save_path.parent.mkdir(parents=True, exist_ok=True)

    output_wb = Workbook()
    default_sheet = output_wb.active
    output_wb.remove(default_sheet)

    errors = []
    sheet_count = 0
    selected_building_name = building_label(selected_building)
    renamed_images = []

    with tempfile.TemporaryDirectory() as temp_dir:
        total = len(input_paths)
        for index, input_path in enumerate(input_paths, start=1):
            def item_progress(value, text, index=index, total=total, name=input_path.name):
                base = ((index - 1) / total) * 100
                scaled = base + (value / total)
                if progress_callback:
                    progress_callback(scaled, f"[{index}/{total}] {name}：{text}")

            try:
                excel_path = prepare_excel_from_input(input_path, temp_dir, item_progress)
                sheet_name = selected_building_name or input_path.stem
                if auto_rename and selected_building_name and input_path.suffix.lower() in IMAGE_EXTENSIONS:
                    renamed_path = rename_image_to_building(input_path, selected_building_name)
                    if renamed_path != input_path:
                        renamed_images.append(f"{input_path.name} → {renamed_path.name}")
                        input_path = renamed_path
                data = extract_arranged_data(excel_path)
                write_arranged_sheet(output_wb, sheet_name, excel_path, data)
                sheet_count += 1
            except Exception as exc:
                errors.append(f"{input_path.name}：{exc}")

    if sheet_count == 0:
        return False, "没有成功生成任何标签页。\n\n" + "\n\n".join(errors[:5])

    if default_output_name(save_path) and selected_building_name:
        save_path = save_path.with_name(f"{selected_building_name}_识别完成.xlsx")

    final_path = make_unique_path(save_path)
    output_wb.save(final_path)
    output_wb.close()

    rename_note = ""
    if renamed_images:
        rename_note = "\n\n已重命名图片：\n" + "\n".join(renamed_images[:20])

    if errors:
        return (
            False,
            f"已生成 {sheet_count} 个标签页，但有 {len(errors)} 个文件失败。\n\n"
            f"保存路径：\n{final_path}{rename_note}\n\n"
            + "\n\n".join(errors[:5]),
        )

    return (
        True,
        f"全部处理完成。\n\n共生成 {sheet_count} 个标签页。\n"
        f"栏位顺序固定为：{', '.join(FIXED_HEADERS)}\n"
        f"左侧保留原表，右侧整理区用公式引用原表，并在下方显示识别数量与核对结果。\n\n"
        f"保存路径：\n{final_path}{rename_note}",
    )


class App(BaseTk):
    def __init__(self):
        super().__init__()
        self.title("称重单识别整理工具")
        self.geometry("930x450")
        self.minsize(850, 410)
        self.resizable(False, False)

        self.input_path = tk.StringVar()
        self.save_path = tk.StringVar()
        self.building = tk.StringVar(value=BUILDING_CHOICES[0])
        self.status = tk.StringVar(value="就绪")
        self.progress = tk.DoubleVar(value=0)
        self.processing = False
        self.input_files = []

        self._build_ui()

    def _build_ui(self):
        root = ttk.Frame(self, padding=24)
        root.pack(fill="both", expand=True)
        root.columnconfigure(1, weight=1)

        ttk.Label(root, text="称重单识别整理工具", font=("Microsoft YaHei UI", 18, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w"
        )

        ttk.Label(
            root,
            text="可拖入图片、Excel 或整个文件夹；每张识别结果一个标签页，左侧原表、右侧整理，并自动核对数量。",
            foreground="#555555",
        ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(6, 22))

        ttk.Label(root, text="栋舍").grid(row=2, column=0, sticky="w", pady=10)
        self.building_combo = ttk.Combobox(
            root,
            textvariable=self.building,
            values=BUILDING_CHOICES,
            state="readonly",
            width=12,
        )
        self.building_combo.grid(row=2, column=1, sticky="w", pady=10)
        self.building_combo.bind("<<ComboboxSelected>>", lambda _event: self.update_default_save_path(force=True))

        ttk.Label(root, text="图片 / Excel / 文件夹").grid(row=3, column=0, sticky="w", pady=10)
        self.input_entry = ttk.Entry(root, textvariable=self.input_path)
        self.input_entry.grid(row=3, column=1, sticky="ew", pady=10)
        self.input_button = ttk.Button(root, text="选择文件", command=self.choose_input)
        self.input_button.grid(row=3, column=2, padx=(12, 0), pady=10)

        ttk.Label(root, text="保存为").grid(row=4, column=0, sticky="w", pady=10)
        self.save_entry = ttk.Entry(root, textvariable=self.save_path)
        self.save_entry.grid(row=4, column=1, sticky="ew", pady=10)
        self.save_button = ttk.Button(root, text="另存为", command=self.choose_save)
        self.save_button.grid(row=4, column=2, padx=(12, 0), pady=10)

        ttk.Label(
            root,
            text="拖入文件夹会自动识别其中所有图片/Excel。图片按所选栋舍重命名，输出表格默认保存到桌面并命名为“栋舍_识别完成.xlsx”。",
            foreground="#666666",
            wraplength=840,
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(16, 8))

        ttk.Label(root, textvariable=self.status, foreground="#333333").grid(
            row=6, column=0, columnspan=3, sticky="ew", pady=(10, 6)
        )
        self.progress_bar = ttk.Progressbar(root, variable=self.progress, maximum=100)
        self.progress_bar.grid(row=7, column=0, columnspan=3, sticky="ew")

        self.run_button = ttk.Button(root, text="开始识别整理", command=self.run)
        self.run_button.grid(row=8, column=2, sticky="e", pady=(24, 0))
        self.update_default_save_path(force=True)
        self._setup_drag_drop()

    def _setup_drag_drop(self):
        if not DND_FILES:
            self.status.set("就绪（拖拽组件不可用，可用“选择文件”多选）")
            return
        self.input_entry.drop_target_register(DND_FILES)
        self.input_entry.dnd_bind("<<Drop>>", self.on_drop)

    def on_drop(self, event):
        self.set_input_files(self.tk.splitlist(event.data))

    def update_default_save_path(self, force=False):
        current = self.save_path.get().strip()
        if current and not force and not default_output_name(current):
            return
        name = building_label(self.building.get()) or "称重单"
        self.save_path.set(str(desktop_path() / f"{name}_识别完成.xlsx"))

    def choose_input(self):
        paths = filedialog.askopenfilenames(
            title="选择称重单图片或 Excel 文件（可多选）",
            filetypes=[
                ("支持的文件", "*.xlsx *.xlsm *.jpg *.jpeg *.png *.bmp *.webp"),
                ("图片文件", "*.jpg *.jpeg *.png *.bmp *.webp"),
                ("Excel 文件", "*.xlsx *.xlsm"),
                ("所有文件", "*.*"),
            ],
        )
        if paths:
            self.set_input_files(paths)
            return
        folder = filedialog.askdirectory(title="或选择一个文件夹")
        if folder:
            self.set_input_files([folder])

    def set_input_files(self, paths):
        if not paths:
            return

        valid_paths = []
        invalid_names = []
        for path in paths:
            file_path = Path(path)
            expanded = expand_input_paths([file_path])
            if expanded:
                valid_paths.extend(expanded)
            else:
                invalid_names.append(file_path.name)

        valid_paths = list(dict.fromkeys(valid_paths))

        if not valid_paths:
            messagebox.showwarning("提示", "请选择或拖入图片、Excel 文件或包含这些文件的文件夹。")
            return

        self.input_files = valid_paths
        if len(valid_paths) == 1:
            self.input_path.set(valid_paths[0])
        else:
            names = [Path(path).name for path in valid_paths[:3]]
            suffix = "" if len(valid_paths) <= 3 else f"；等 {len(valid_paths)} 个文件"
            self.input_path.set("；".join(names) + suffix)

        self.update_default_save_path()

        self.status.set(f"已选择 {len(valid_paths)} 个文件，将保存到同一个 Excel。")
        if invalid_names:
            messagebox.showwarning("提示", "以下文件已忽略：\n" + "\n".join(invalid_names[:10]))

    def choose_save(self):
        path = filedialog.asksaveasfilename(
            title="保存识别整理结果",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self.save_path.set(path)

    def run(self):
        if self.processing:
            return

        input_paths = list(self.input_files)
        save_path = self.save_path.get().strip()
        selected_building = self.building.get().strip()
        if not input_paths:
            messagebox.showwarning("提示", "请先选择图片或 Excel 文件。")
            return
        if selected_building not in BUILDING_CHOICES:
            messagebox.showwarning("提示", "请先选择栋舍。")
            return
        for input_path in input_paths:
            if not Path(input_path).exists():
                messagebox.showwarning("提示", f"输入文件不存在：\n{input_path}")
                return
            if Path(input_path).suffix.lower() not in SUPPORTED_EXTENSIONS:
                messagebox.showwarning("提示", "请选择图片文件、.xlsx/.xlsm 文件，或拖入文件夹。")
                return
        if not save_path:
            messagebox.showwarning("提示", "请先选择保存路径。")
            return

        self.processing = True
        self._set_controls_state(tk.DISABLED)
        self.progress.set(8)
        self.status.set("正在准备...")
        threading.Thread(
            target=self._worker,
            args=(input_paths, save_path, selected_building),
            daemon=True,
        ).start()

    def _worker(self, input_paths, save_path, selected_building):
        ok, msg = process_inputs_to_workbook(
            input_paths,
            save_path,
            self._set_progress_threadsafe,
            selected_building=selected_building,
        )
        self.after(0, lambda: self._finish(ok, msg))

    def _finish(self, ok, msg):
        self.processing = False
        self._set_controls_state(tk.NORMAL)
        self.progress.set(100 if ok else 0)
        self.status.set("处理完成" if ok else "处理失败")
        if ok:
            messagebox.showinfo("处理完成", msg)
        else:
            messagebox.showwarning("处理结果", msg)

    def _set_progress_threadsafe(self, value, text):
        def update():
            self.progress.set(max(0, min(100, value)))
            self.status.set(text)

        self.after(0, update)

    def _set_controls_state(self, state):
        for widget in (
            self.input_entry,
            self.input_button,
            self.building_combo,
            self.save_entry,
            self.save_button,
            self.run_button,
        ):
            widget.config(state=state)


if __name__ == "__main__":
    App().mainloop()
